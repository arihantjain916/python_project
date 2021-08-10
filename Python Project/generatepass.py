def genpass():
    import check_pass as cp
    n = 0
    while n == 0:
        import openpyxl
        import random
        import pandas as pd

        max_len = 12

        a = input("Enter the username: ").lower()

        df1 = pd.DataFrame(pd.read_excel(r"Book1.xlsx"))# , sheet_name = "Sheet1"))
        data = df1[['username']]
        user=[]    
        for i in range(0,len(data)):
            user.append(data.loc[i][0])
        if(a in user):
            for x,y in enumerate(user):
                if(a==y):
                    # print(df1.iloc[x,:3])
                    print("Username already exist") 
                    n = 0
                    v = input("You want to login via your username or try again with different username(login/try again): ").lower()
                    if v == 'login':
                        cp.fetch_user()
                        n = 1
                    elif v == 'try again':
                        continue
                    else:
                        print("Invalid entry")
                        exit()
        else:          
            digits = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']

            lower = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h',
                                'i', 'j', 'k', 'm', 'n', 'o', 'p', 'q',
                                'r', 's', 't', 'u', 'v', 'w', 'x', 'y',
                                'z']

            upper = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H',
                                'I', 'J', 'K', 'M', 'N', 'O', 'p', 'Q',
                                'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
                                'Z']

            symbol = ['@', '#', '$', '%', '=', ':', '?', '.', '/', '|', '~', '>',
                    '*', '(', ')', '<']


            combined = digits + upper + lower + symbol


            rand_digit = random.choice(digits)
            rand_upper = random.choice(upper)
            rand_lower = random.choice(lower)
            rand_symbol = random.choice(symbol)


            temp_pass = rand_digit + rand_upper + rand_lower + rand_symbol



            for x in range(max_len - 4):
                temp_pass = temp_pass + random.choice(combined)
                temp_pass_list = list(temp_pass)
                random.shuffle(temp_pass_list)


            password = ""
            for x in temp_pass_list:
                    password = password + x
                    
            # print out password
            # print(password)
            workbook =openpyxl.load_workbook("Book1.xlsx")
            worksheet = workbook["Sheet1"]
            rows_old = worksheet.max_row

            data=[]
            data=[a , password]
            for j in range(1,3):
                c3 = worksheet.cell(row = rows_old+1, column = j)
                c3.value = data[j-1]
            workbook.save("Book1.xlsx")
            n += 1
            print("Thank you for sign up...")
            print("This is your login credentials")
            print(f"Your username is {a} and your password is {password}")
            print("\n\n Please login using your login credentials")
# genpass()